"""
Geração de relatórios operacionais exportáveis: PDF executivo (ReportLab) e
planilha Excel multi-aba (Pandas + OpenPyXL).

Ambos os formatos partem dos mesmos dados brutos (`coletar_dados`), lidos de
`metricas_atendimento` / `metricas_ocupacao` no intervalo [data_inicio, data_fim].
"""
import io
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time as dt_time, timedelta
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

import models

# Turnos usados no "resumo por dia/turno" — ajuste aqui se a operação usar outra convenção.
_TURNOS = [
    ("Madrugada", 0, 6),
    ("Manhã", 6, 12),
    ("Tarde", 12, 18),
    ("Noite", 18, 24),
]

_COR_PRIMARIA = colors.HexColor("#0f172a")
_COR_TEXTO_SUAVE = colors.HexColor("#475569")
_COR_LINHA = colors.HexColor("#e2e8f0")
_COR_ZEBRA = colors.HexColor("#f8fafc")
_COR_DESTAQUE = colors.HexColor("#0ea5e9")


def _turno_da_hora(hora: int) -> str:
    for nome, inicio, fim in _TURNOS:
        if inicio <= hora < fim:
            return nome
    return _TURNOS[0][0]


def _slug(texto: str) -> str:
    """Converte um nome livre (com acentos/espaços) em algo seguro para nome de arquivo."""
    normalizado = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", normalizado).strip("-").lower()
    return slug or "relatorio"


def _fmt_duracao(segundos: float) -> str:
    minutos, s = divmod(int(round(max(0, segundos))), 60)
    return f"{minutos}m {s}s"


# ==============================================================================
# Coleta de dados (compartilhada entre PDF e Excel)
# ==============================================================================
@dataclass
class DadosRelatorio:
    empresa: models.Empresa
    data_inicio: date
    data_fim: date
    atendimentos: List[models.MetricaAtendimento]
    ocupacoes: List[models.MetricaOcupacao]
    nomes_cameras: Dict[int, str]


def coletar_dados(db: Session, empresa: models.Empresa, data_inicio: date, data_fim: date) -> DadosRelatorio:
    inicio_dt = datetime.combine(data_inicio, dt_time.min)
    fim_dt = datetime.combine(data_fim, dt_time.max)

    atendimentos = (
        db.query(models.MetricaAtendimento)
        .filter(
            models.MetricaAtendimento.empresa_id == empresa.id,
            models.MetricaAtendimento.timestamp.between(inicio_dt, fim_dt),
        )
        .order_by(models.MetricaAtendimento.timestamp)
        .all()
    )
    ocupacoes = (
        db.query(models.MetricaOcupacao)
        .filter(
            models.MetricaOcupacao.empresa_id == empresa.id,
            models.MetricaOcupacao.timestamp.between(inicio_dt, fim_dt),
        )
        .order_by(models.MetricaOcupacao.timestamp)
        .all()
    )
    cameras = db.query(models.Camera).filter(models.Camera.empresa_id == empresa.id).all()
    nomes_cameras = {c.id: c.nome_camera for c in cameras}

    return DadosRelatorio(empresa, data_inicio, data_fim, atendimentos, ocupacoes, nomes_cameras)


def nome_arquivo(dados: DadosRelatorio, extensao: str) -> str:
    return (
        f"relatorio-{_slug(dados.empresa.nome_empresa)}-"
        f"{dados.data_inicio.isoformat()}-a-{dados.data_fim.isoformat()}.{extensao}"
    )


def _resumo_diario_turno(dados: DadosRelatorio) -> List[dict]:
    grupos = defaultdict(list)
    for a in dados.atendimentos:
        grupos[(a.timestamp.date(), _turno_da_hora(a.timestamp.hour))].append(a)

    linhas = []
    for (dia, turno), itens in sorted(grupos.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        total = len(itens)
        concluidos = sum(1 for i in itens if i.concluido)
        tempo_medio = sum(i.duracao_segundos for i in itens) / total if total else 0.0
        linhas.append(
            {
                "data": dia,
                "turno": turno,
                "total_atendimentos": total,
                "concluidos": concluidos,
                "abandonados": total - concluidos,
                "tempo_medio_segundos": round(tempo_medio, 1),
            }
        )
    return linhas


def _horarios_pico(dados: DadosRelatorio) -> Dict[int, int]:
    contagem: Dict[int, int] = defaultdict(int)
    for a in dados.atendimentos:
        contagem[a.timestamp.hour] += 1
    return contagem


# ==============================================================================
# PDF (ReportLab) — relatório executivo
# ==============================================================================
def _grafico_horarios_pico(horarios: Dict[int, int]) -> Drawing:
    valores = [horarios.get(h, 0) for h in range(24)]
    desenho = Drawing(460, 150)

    grafico = VerticalBarChart()
    grafico.x, grafico.y = 30, 22
    grafico.width, grafico.height = 415, 110
    grafico.data = [valores]
    grafico.categoryAxis.categoryNames = [f"{h}h" for h in range(24)]
    grafico.categoryAxis.labels.fontSize = 5.5
    grafico.categoryAxis.labels.angle = 90
    grafico.categoryAxis.labels.dy = -14
    grafico.valueAxis.valueMin = 0
    grafico.valueAxis.labels.fontSize = 6.5
    grafico.bars[0].fillColor = _COR_DESTAQUE
    grafico.barWidth = 5
    grafico.groupSpacing = 2

    desenho.add(grafico)
    return desenho


def gerar_pdf(dados: DadosRelatorio) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=1.6 * cm,
        bottomMargin=1.4 * cm,
        leftMargin=1.8 * cm,
        rightMargin=1.8 * cm,
        title=f"Relatorio Operacional - {dados.empresa.nome_empresa}",
    )

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloEmpresa", parent=estilos["Title"], fontSize=18, alignment=0, textColor=_COR_PRIMARIA, spaceAfter=2
    )
    estilo_subtitulo = ParagraphStyle(
        "Subtitulo", parent=estilos["Normal"], fontSize=10.5, textColor=_COR_TEXTO_SUAVE, spaceAfter=2
    )
    estilo_secao = ParagraphStyle(
        "Secao", parent=estilos["Heading2"], fontSize=13, textColor=_COR_PRIMARIA, spaceBefore=16, spaceAfter=8
    )
    estilo_normal = estilos["Normal"]
    estilo_nota = ParagraphStyle(
        "Nota", parent=estilos["Normal"], fontSize=8.5, textColor=_COR_TEXTO_SUAVE, leading=12
    )

    elementos = []

    # --- Cabeçalho da empresa ---
    elementos.append(Paragraph(dados.empresa.nome_empresa, estilo_titulo))
    elementos.append(Paragraph("Relatório Operacional — Inteligência por Câmeras", estilo_subtitulo))
    elementos.append(
        Paragraph(
            f"Período: {dados.data_inicio.strftime('%d/%m/%Y')} a {dados.data_fim.strftime('%d/%m/%Y')}",
            estilo_subtitulo,
        )
    )
    elementos.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", estilo_subtitulo))
    elementos.append(Spacer(1, 0.5 * cm))

    # --- KPIs ---
    total = len(dados.atendimentos)
    concluidos = sum(1 for a in dados.atendimentos if a.concluido)
    tempo_medio = sum(a.duracao_segundos for a in dados.atendimentos) / total if total else 0.0
    taxa = (concluidos / total * 100) if total else 0.0

    tabela_kpi = Table(
        [
            ["Total de Atendimentos", "Tempo Médio por Atendimento", "Taxa de Conclusão"],
            [str(total), _fmt_duracao(tempo_medio), f"{taxa:.1f}%"],
        ],
        colWidths=[5.6 * cm] * 3,
    )
    tabela_kpi.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _COR_PRIMARIA),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (-1, 1), 17),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 1), (-1, 1), 10),
                ("BOTTOMPADDING", (0, 1), (-1, 1), 12),
                ("BACKGROUND", (0, 1), (-1, 1), _COR_ZEBRA),
                ("GRID", (0, 0), (-1, -1), 0.5, _COR_LINHA),
            ]
        )
    )
    elementos.append(tabela_kpi)

    # --- Resumo por dia/turno ---
    elementos.append(Paragraph("Resumo por Dia e Turno", estilo_secao))
    resumo = _resumo_diario_turno(dados)
    if resumo:
        linhas = [["Data", "Turno", "Atendimentos", "Concluídos", "Abandonados", "Tempo Médio"]]
        for r in resumo:
            linhas.append(
                [
                    r["data"].strftime("%d/%m/%Y"),
                    r["turno"],
                    str(r["total_atendimentos"]),
                    str(r["concluidos"]),
                    str(r["abandonados"]),
                    _fmt_duracao(r["tempo_medio_segundos"]),
                ]
            )
        tabela = Table(
            linhas, colWidths=[2.6 * cm, 2.6 * cm, 2.8 * cm, 2.6 * cm, 2.8 * cm, 2.8 * cm], repeatRows=1
        )
        tabela.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), _COR_PRIMARIA),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("ALIGN", (2, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.4, _COR_LINHA),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _COR_ZEBRA]),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        elementos.append(tabela)
    else:
        elementos.append(Paragraph("Nenhum atendimento registrado no período selecionado.", estilo_normal))

    # --- Gráfico de horários de pico ---
    elementos.append(Paragraph("Horários de Pico (atendimentos por hora)", estilo_secao))
    horarios = _horarios_pico(dados)
    if horarios:
        elementos.append(_grafico_horarios_pico(horarios))
    else:
        elementos.append(Paragraph("Sem dados suficientes para o gráfico neste período.", estilo_normal))

    # --- Nota de conformidade LGPD ---
    elementos.append(Paragraph("Conformidade LGPD", estilo_secao))
    elementos.append(
        Paragraph(
            "Este relatório contém apenas métricas agregadas (contagens e durações). Nenhuma imagem, "
            "recorte facial ou dado biométrico é armazenado pelo sistema: todo frame de vídeo processado "
            "tem rostos e pessoas automaticamente anonimizados (Gaussian Blur) antes de qualquer exibição "
            "ou análise, em conformidade com a Lei Geral de Proteção de Dados (Lei nº 13.709/2018).",
            estilo_nota,
        )
    )

    doc.build(elementos)
    buffer.seek(0)
    return buffer.getvalue()


# ==============================================================================
# Excel (Pandas + OpenPyXL) — planilha multi-aba
# ==============================================================================
def _formatar_aba(ws) -> None:
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    preenchimento = PatternFill("solid", fgColor="0F172A")
    for celula in ws[1]:
        celula.font = fonte_cabecalho
        celula.fill = preenchimento
    ws.freeze_panes = "A2"
    for coluna_celulas in ws.columns:
        comprimento = max((len(str(c.value)) for c in coluna_celulas if c.value is not None), default=8)
        letra = get_column_letter(coluna_celulas[0].column)
        ws.column_dimensions[letra].width = min(42, comprimento + 3)


def gerar_excel(dados: DadosRelatorio) -> bytes:
    # --- Aba "Resumo Diário" ---
    resumo = _resumo_diario_turno(dados)
    colunas_resumo = ["Data", "Turno", "Total de Atendimentos", "Concluídos", "Abandonados", "Tempo Médio (s)"]
    df_resumo = pd.DataFrame(
        [
            {
                "Data": r["data"].strftime("%d/%m/%Y"),
                "Turno": r["turno"],
                "Total de Atendimentos": r["total_atendimentos"],
                "Concluídos": r["concluidos"],
                "Abandonados": r["abandonados"],
                "Tempo Médio (s)": r["tempo_medio_segundos"],
            }
            for r in resumo
        ],
        columns=colunas_resumo,
    )

    # --- Aba "Log de Atendimentos" ---
    # A tabela só grava o timestamp de ENCERRAMENTO + a duração; a hora de início é
    # derivada (timestamp - duração) e documentada aqui para não confundir quem ler.
    colunas_log = ["Data", "Câmera", "Hora Início", "Hora Fim", "Duração", "Duração (s)", "Concluído"]
    linhas_log = []
    for a in dados.atendimentos:
        hora_fim = a.timestamp
        hora_inicio = hora_fim - timedelta(seconds=a.duracao_segundos)
        linhas_log.append(
            {
                "Data": hora_fim.strftime("%d/%m/%Y"),
                "Câmera": dados.nomes_cameras.get(a.camera_id, f"Câmera #{a.camera_id}"),
                "Hora Início": hora_inicio.strftime("%H:%M:%S"),
                "Hora Fim": hora_fim.strftime("%H:%M:%S"),
                "Duração": _fmt_duracao(a.duracao_segundos),
                "Duração (s)": round(a.duracao_segundos, 1),
                "Concluído": "Sim" if a.concluido else "Não",
            }
        )
    df_log = pd.DataFrame(linhas_log, columns=colunas_log)

    # --- Aba "Métricas de Ocupação" ---
    colunas_ocup = ["Data", "Hora", "Câmera", "Pessoas Detectadas", "Tempo de Inatividade (s)"]
    df_ocup = pd.DataFrame(
        [
            {
                "Data": o.timestamp.strftime("%d/%m/%Y"),
                "Hora": o.timestamp.strftime("%H:%M:%S"),
                "Câmera": dados.nomes_cameras.get(o.camera_id, f"Câmera #{o.camera_id}"),
                "Pessoas Detectadas": o.pessoas_detectadas,
                "Tempo de Inatividade (s)": o.tempo_inatividade_segundos,
            }
            for o in dados.ocupacoes
        ],
        columns=colunas_ocup,
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo Diário", index=False)
        df_log.to_excel(writer, sheet_name="Log de Atendimentos", index=False)
        df_ocup.to_excel(writer, sheet_name="Métricas de Ocupação", index=False)
        for nome_aba in writer.sheets:
            _formatar_aba(writer.sheets[nome_aba])

    buffer.seek(0)
    return buffer.getvalue()
